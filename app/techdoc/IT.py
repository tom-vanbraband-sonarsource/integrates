# -*- coding: utf-8 -*-
""" Class for generate an xlsx file with findings information. """
import re
from openpyxl import load_workbook


class ITReport(object):
    """Class to generate IT reports."""

    workbook = None
    current_sheet = None
    data = None
    lang = None
    row = 3
    result_filename = ''
    result_path = '/usr/src/app/app/techdoc/results/'
    templates = {
        'es': {
            'TECHNICAL': '/usr/src/app/app/techdoc/templates/TECHNICAL.xlsx',
        },
        'en': {}}
    sheet_names = {
        'es': {
            'finding': 'Findings',
        },
        'en': {}}
    finding = {
        'name': 2,
        'description': 3,
        'where': 4,
        'where_records': 5,
        'measurements': 7,
        'severityCvss': 8,
        'cardinality': 9,
        'affected_records': 10,
        'evidence': 11,
        'solution': 12,
        'requirements_id': 13}
    matriz = {
        'type': 5,
        'component': 6,
        'requirements_id': 8,
        'requirements': 9,
        'scenario': 13,
        'ambit': 17,
        'category': 18,
        'threat': 19,
        'cssv3_value': 24,
        'probability': 26,
        'severity': 27,
        'risk': 30}

    def __init__(self, project, data, username, lang='es'):
        """Initialize variables."""
        self.lang = lang
        self.workbook = load_workbook(
            filename=self.templates[self.lang]['TECHNICAL']
        )
        self.generate(data, project, username)

    def hide_cell(self, data):
        init_row = 3 + 12 * len(data)
        end_row = 3 + 12 * 60
        self.__select_finding_sheet()
        for row in range(init_row, end_row):
            self.current_sheet.row_dimensions[row].hidden = True

    def generate(self, data, project, username):
        for finding in data:
            self.__write(finding)
            self.row += 12
        self.hide_cell(data)
        self.__save(project, username)

    def __select_finding_sheet(self):
        """Select finding sheet."""
        self.current_sheet = self.workbook[
            self.sheet_names[self.lang]['finding']
        ]

    def set_cell(self, col, value, inc=0):
        """Assign a value to a cell with findings index."""
        self.current_sheet.cell(row=self.row + inc, column=col).value = value

    def set_cell_number(self, col, value, inc=0):
        """Assign a numeric value to a cell with findings index."""
        self.current_sheet.cell(row=self.row + inc, column=col).value = float(value)

    def __get_req(self, req_vect): # noqa
        """Get all the identifiers with the REQ.XXXX format."""
        try:
            reqs = re.findall('REQ\\.\\d{3,4}', req_vect) # noqa
            reqs = [x.replace('REQ.', '') for x in reqs]
            reqs_list = '|'.join(reqs)
            return '.*(' + reqs_list + ')'
        except ValueError:
            return ''

    def __get_measure(self, metric, metric_value): # noqa
        """Extract number of CSSV metrics."""
        try:
            metrics = {
                'attackVector': {
                    '0.85': 'Network',
                    '0.62': 'Adjacent',
                    '0.55': 'Local',
                    '0.20': 'Physical',
                },
                'attackComplexity': {
                    '0.77': 'Low',
                    '0.44': 'High',
                },
                'privilegesRequired': {
                    '0.85': 'None',
                    '0.62': 'Low',
                    '0.68': 'Low',
                    '0.27': 'High',
                    '0.50': 'High',
                },
                'userInteraction': {
                    '0.85': 'None',
                    '0.62': 'Required',
                },
                'severityScope': {
                    '0.0': 'Unchanged',
                    '1.0': 'Changed',
                },
                'confidentialityImpact': {
                    '0.56': 'High',
                    '0.22': 'low',
                    '0.0': 'None',
                },
                'integrityImpact': {
                    '0.56': 'High',
                    '0.22': 'Low',
                    '0.0': 'None',
                },
                'availabilityImpact': {
                    '0.56': 'High',
                    '0.22': 'Low',
                    '0.0': 'None',
                },
                'exploitability': {
                    '0.91': 'Unproven',
                    '0.94': 'Proof of concept',
                    '0.97': 'Functional',
                    '1.0': 'High',
                },
                'remediationLevel': {
                    '0.95': 'Oficial Fix',
                    '0.96': 'Temporary Fix',
                    '0.97': 'Workaround',
                    '1.0': 'Unavailable',
                },
                'reportConfidence': {
                    '0.92': 'Unknown',
                    '0.96': 'Reasonable',
                    '1.0': 'Confirmed',
                }
            }
            metric_descriptions = metrics.get(metric)
            if metric_descriptions:
                description = metric_descriptions.get(str(metric_value))
            else:
                description = ''
            return description
        except ValueError:
            return ''

    def __write(self, row):
        """Write Formstack finding in a row on the Findings sheet."""
        self.__select_finding_sheet()
        self.set_cell(self.finding['name'], row['finding'])
        self.set_cell(self.finding['description'], row['vulnerability'])
        self.set_cell(self.finding['where'], row['where'])
        if int(row['recordsNumber']) != 0:
            self.set_cell(self.finding['where_records'],
                          'Evidences/' + row['finding'] + '/records.csv')
        self.set_cell_number(self.finding['severityCvss'], row['severityCvss'])
        self.set_cell_number(self.finding['cardinality'], row['openVulnerabilities'])
        self.set_cell_number(self.finding['affected_records'], row['recordsNumber'])
        self.set_cell(self.finding['evidence'], 'Evidences/' + row['finding'])
        self.set_cell(self.finding['solution'], row['effectSolution'])
        self.set_cell(self.finding['requirements_id'],
                      self.__get_req(row['requirements']))
        self.set_cell(
            self.finding['measurements'],
            self.__get_measure('attackVector', row['attackVector']))
        self.set_cell(
            self.finding['measurements'],
            self.__get_measure('attackComplexity', row['attackComplexity']), 1)
        self.set_cell(
            self.finding['measurements'],
            self.__get_measure(
                'privilegesRequired', row['privilegesRequired']),
            2)
        self.set_cell(
            self.finding['measurements'],
            self.__get_measure('userInteraction', row['userInteraction']), 3)
        self.set_cell(
            self.finding['measurements'],
            self.__get_measure('severityScope', row['severityScope']), 4)
        self.set_cell(
            self.finding['measurements'],
            self.__get_measure(
                'confidentialityImpact', row['confidentialityImpact']),
            5)
        self.set_cell(
            self.finding['measurements'],
            self.__get_measure('integrityImpact', row['integrityImpact']), 6)
        self.set_cell(
            self.finding['measurements'],
            self.__get_measure(
                'availabilityImpact', row['availabilityImpact']),
            7)
        self.set_cell(
            self.finding['measurements'],
            self.__get_measure('exploitability', row['exploitability']), 8)
        self.set_cell(
            self.finding['measurements'],
            self.__get_measure('remediationLevel', row['remediationLevel']), 9)
        self.set_cell(
            self.finding['measurements'],
            self.__get_measure('reportConfidence', row['reportConfidence']), 10)

    def __save(self, project, username):
        self.result_filename = self.result_path
        self.result_filename += project + '_' + username + '.xlsx'
        self.workbook.save(self.result_filename)


def translate_parameter(param):
    translation_values = {
        'CONTINUOUS': 'Continua',
        'ANALYSIS': 'Análisis',
        'APP': 'Aplicación',
        'BINARY': 'Binario',
        'SOURCE_CODE': 'Código fuente',
        'INFRASTRUCTURE': 'Infraestructura',
        'ANONYMOUS_INTERNET': 'Anónimo desde internet',
        'ANONYMOUS_INTRANET': 'Anónimo desde intranet',
        'AUTHORIZED_USER_EXTRANET': 'Extranet usuario autorizado',
        'UNAUTHORIZED_USER_EXTRANET': 'Extranet usuario no autorizado',
        'AUTHORIZED_USER_INTERNET': 'Internet usuario autorizado',
        'UNAUTHORIZED_USER_INTERNET': 'Internet usuario no autorizado',
        'AUTHORIZED_USER_INTRANET': 'Intranet usuario autorizado',
        'UNAUTHORIZED_USER_INTRANET': 'Intranet usuario no autorizado',
        'APPLICATIONS': 'Aplicaciones',
        'DATABASES': 'Bases de Datos'
    }
    return translation_values.get(param)


def get_probability(probability):
    probability_values = {
        '100': '100% Vulnerado Anteriormente',
        '75': '75% Fácil de vulnerar',
        '50': '50% Posible de vulnerar',
        '25': '25% Difícil de vulnerar'
    }
    return probability_values.get(str(probability))


def cast_severity(severity):
    """Cast severity value."""
    severity_value = ''
    if severity >= 9.0 and severity <= 10.0:
        severity_value = 'Crítica'
    elif severity >= 7.0 and severity <= 8.9:
        severity_value = 'Alta'
    elif severity >= 4.0 and severity <= 6.9:
        severity_value = 'Media'
    elif severity >= 0.1 and severity <= 3.9:
        severity_value = 'Baja'
    else:
        severity_value = 'Ninguna'
    return severity_value
