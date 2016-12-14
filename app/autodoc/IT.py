""" 
    Modulo del AutodocIntegrates para generar
    las presentaciones automaticas xlsx de un
    proyecto en:
    - Fluid 
    - Bancolombia
"""
from openpyxl import load_workbook
import re

class Bancolombia:

    workbook = None
    template_path = "/var/www/fluid-integrates/app/autodoc/templates/bancolombia.xlsx"
    result_path = "/var/www/fluid-integrates/app/autodoc/results/:project.xlsx"
    current_sheet = None 
    project_name = ""
    project_data = None
    COL_HALLAZGO = 2
    COL_VULNERABILIDAD = 3
    COL_DONDE = 4
    COL_REQUISITOS = 5
    COL_METRICAS = 7
    COL_CRITICIDAD = 8
    COL_CARDINALIDAD = 9
    COL_EVIDENCIA = 10
    COL_SOLUCION = 11
    COL_INSTRUCTIVO = 12
    COL_ID_REQUISITOS = 13
    COL_AMENAZA = 14
    COL_RIESGO = 15
    #CONSTANTES QC
    COLQC_TIPO_PRUEBA = 5
    COLQC_COMPONENTE_APLICATIVO = 6
    COLQC_ID_REQUISITOS = 8
    COLQC_FABRICA = 10
    COLQC_DETECTADO_POR = 11
    COLQC_DETECTADO_CICLO = 12
    COLQC_ESCENARIO = 13
    COLQC_HALLAZGO = 14
    COLQC_AMBITO = 17
    COLQC_CATEGORIA = 18
    COLQC_AMENAZA = 19
    COLQC_TIPO_REPORTE = 20
    COLQC_NATURALEZA = 21
    COLQC_PROBABILIDAD = 26
    COLQC_SEVERIDAD = 27
    row = 3
    qc_row = 3

    def __init__(self, project, data):
        self.workbook = load_workbook(filename=self.template_path)
        self.project_name = project 
        self.project_data = data
        self.generate_doc()
        self.save()

    def fill_finding(self, finding):
        self.select_finding_sheet()
        self.assign(self.COL_HALLAZGO, finding["hallazgo"])
        self.assign(self.COL_VULNERABILIDAD, finding["vulnerabilidad"])
        self.assign(self.COL_DONDE, finding["donde"])
        self.assign(self.COL_REQUISITOS, finding["requisitos"])
        self.assign(self.COL_CRITICIDAD, finding["criticidad"])
        self.assign(self.COL_CARDINALIDAD, finding["cardinalidad"])
        self.assign(self.COL_EVIDENCIA, "Evidencias/" + finding["hallazgo"])
        self.assign(self.COL_SOLUCION, finding["solucion_efecto"])
        self.assign(self.COL_INSTRUCTIVO, "Soluciones/"+ finding["hallazgo"])
        self.assign(self.COL_ID_REQUISITOS, self.extract_reqs(finding["requisitos"]))
        self.assign(self.COL_AMENAZA, finding["amenaza"])
        self.assign(self.COL_RIESGO, finding["riesgo"])
        #Metricas
        self.assign(self.COL_METRICAS, self.extract_metric(finding["vector_acceso"]))
        self.assign(self.COL_METRICAS, self.fix_complexity_access(self.extract_metric(finding["complejidad_acceso"])), 1)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["autenticacion"]), 2)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["impacto_confidencialidad"]), 3)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["impacto_integridad"]), 4)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["impacto_disponibilidad"]), 5)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["explotabilidad"]), 6)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["nivel_resolucion"]), 7)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["nivel_confianza"]), 8)

    def fill_qc(self, finding):
        self.select_qc_sheet()
        self.assign_qc(self.COLQC_TIPO_PRUEBA, finding["tipo_prueba"])
        self.assign_qc(self.COLQC_ID_REQUISITOS, self.extract_reqs(finding["requisitos"]))
        self.assign_qc(self.COLQC_FABRICA, "Fluid")
        self.assign_qc(self.COLQC_DETECTADO_POR, finding["analista"])
        if "escenario" in finding:
            self.assign_qc(self.ESCENARIO, finding["escenario"])
        self.assign_qc(self.COLQC_HALLAZGO, finding["hallazgo"])
        self.assign_qc(self.COLQC_AMBITO, finding["ambito"])
        if "categoria" in finding:
            self.assign_qc(self.COLQC_CATEGORIA, finding["categoria"])
        self.assign_qc(self.COLQC_AMENAZA, finding["amenaza"])
        self.assign_qc(self.COLQC_TIPO_REPORTE, "Hallazgo")
        self.assign_qc(self.COLQC_NATURALEZA, "Seguridad")
        self.assign_qc(self.COLQC_PROBABILIDAD, finding["probabilidad"])
        self.assign_qc(self.COLQC_SEVERIDAD, finding["severidad"])

    def fix_complexity_access(self, complexity_access):
        if complexity_access == "Bajo":
            return "Baja"
        elif complexity_access == "Alto":
            return "Alta"
        elif complexity_access == "Medio":
            return "Media"
        else: 
            return ""

    def generate_doc(self):
        for finding in self.project_data:
            self.fill_finding(finding)
            self.fill_qc(finding)
            self.row += 10
            self.qc_row += 1

    def select_finding_sheet(self):
            self.current_sheet = self.workbook["Hallazgos"]

    def select_qc_sheet(self):
            self.current_sheet = self.workbook["MatrizQC"]
            
    def assign(self, col, value, inc = 0):
        """ Asigna un valor a una celda """
        self.current_sheet.cell(row= self.row + inc, column=col).value = value
    
    def assign_qc(self, col, value, inc = 0):
        """ Asigna un valor a una celda """
        self.current_sheet.cell(row= self.qc_row + inc, column=col).value = value

    def extract_reqs(self, req_vect):
        """ Obtiene todos los identificadores con el formato REQ.XXXX """
        try:
            reqs = re.findall("REQ\\.\\d{3,4}", req_vect)
            reqs = [x.replace("REQ.", "") for x in reqs]
            reqs_list = "|".join(reqs)
            return ".*(" + reqs_list + ")"
        except ValueError:
            return ""

    def extract_metric(self, metric_str):
        "Obtiene los valores de la cadena de texto con las \
        metricas de calificacion de formstack"
        try:
            return metric_str.split("|")[1].strip().split(":")[0].strip()
        except ValueError:
            return ""

    def save(self):
        self.workbook.save(self.result_path.replace(":project",self.project_name))

class Fluid:

    workbook = None
    template_path = "/var/www/fluid-integrates/app/autodoc/templates/fluid.xlsx"
    result_path = "/var/www/fluid-integrates/app/autodoc/results/:project.xlsx"
    current_sheet = None 
    project_name = ""
    project_data = None
    COL_HALLAZGO = 2
    COL_VULNERABILIDAD = 3
    COL_DONDE = 4
    COL_REQUISITOS = 5
    COL_METRICAS = 7
    COL_CRITICIDAD = 8
    COL_CARDINALIDAD = 9
    COL_EVIDENCIA = 10
    COL_SOLUCION = 11
    COL_INSTRUCTIVO = 12
    COL_ID_REQUISITOS = 13
    row = 3

    def __init__(self, project, data):
        self.workbook = load_workbook(filename=self.template_path)
        self.project_name = project 
        self.project_data = data
        self.generate_doc()
        self.save()

    def fill_finding(self, finding):
        self.select_finding_sheet()
        self.assign(self.COL_HALLAZGO, finding["hallazgo"])
        self.assign(self.COL_VULNERABILIDAD, finding["vulnerabilidad"])
        self.assign(self.COL_DONDE, finding["donde"])
        self.assign(self.COL_REQUISITOS, finding["requisitos"])
        self.assign(self.COL_CRITICIDAD, finding["criticidad"])
        self.assign(self.COL_CARDINALIDAD, finding["cardinalidad"])
        self.assign(self.COL_EVIDENCIA, "Evidencias/" + finding["hallazgo"])
        self.assign(self.COL_SOLUCION, finding["solucion_efecto"])
        self.assign(self.COL_INSTRUCTIVO, "Soluciones/"+ finding["hallazgo"])
        self.assign(self.COL_ID_REQUISITOS, self.extract_reqs(finding["requisitos"]))
        #Metricas
        self.assign(self.COL_METRICAS, self.extract_metric(finding["vector_acceso"]))
        self.assign(self.COL_METRICAS, self.fix_complexity_access(self.extract_metric(finding["complejidad_acceso"])), 1)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["autenticacion"]), 2)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["impacto_confidencialidad"]), 3)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["impacto_integridad"]), 4)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["impacto_disponibilidad"]), 5)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["explotabilidad"]), 6)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["nivel_resolucion"]), 7)
        self.assign(self.COL_METRICAS, self.extract_metric(finding["nivel_confianza"]), 8)

    def fix_complexity_access(self, complexity_access):
        if complexity_access == "Bajo":
            return "Baja"
        elif complexity_access == "Alto":
            return "Alta"
        elif complexity_access == "Medio":
            return "Media"
        else: 
            return ""

    def generate_doc(self):
        for finding in self.project_data:
            self.fill_finding(finding)
            self.row += 10

    def select_finding_sheet(self):
            self.current_sheet = self.workbook["Hallazgos"]
            
    def assign(self, col, value, inc = 0):
        """ Asigna un valor a una celda """
        self.current_sheet.cell(row=self.row + inc, column=col).value = value

    def extract_reqs(self, req_vect):
        """ Obtiene todos los identificadores con el formato REQ.XXXX """
        try:
            reqs = re.findall("REQ\\.\\d{3,4}", req_vect)
            reqs = [x.replace("REQ.", "") for x in reqs]
            reqs_list = "|".join(reqs)
            return ".*(" + reqs_list + ")"
        except ValueError:
            return ""

    def extract_metric(self, metric_str):
        "Obtiene los valores de la cadena de texto con las \
        metricas de calificacion de formstack"
        try:
            return metric_str.split("|")[1].strip().split(":")[0].strip()
        except ValueError:
            return ""

    def save(self):
        self.workbook.save(self.result_path.replace(":project",self.project_name))