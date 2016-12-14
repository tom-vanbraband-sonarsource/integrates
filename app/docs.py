# -*- coding: utf-8 -*-
"""
	Crear informe tecnico a partir de un JSON y guardarlo en formato excel
	utilizando openpyxl
	TODO: Verificar como arreglar openpyxl hay un error cuando los nombres
    de los sheet exceden 31 caracteres
"""
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader
from os import path
import codecs, util, re

config = {
    'dir': path.dirname(path.dirname(path.abspath(__file__))),
    'trim_blocks': True,
    'jinja': {
        'general': 'templates/general.txt',
        'detallado': 'templates/detallado.txt'
    },
    'xls': {
        'general': '/app/templates/fluid.xlsx',
        'detallado': '/app/templates/bancolombia.xlsx',
        'first_row': 3,
    }
}

def create_template(project_name, project_data):
    """Convierte un json en un template de jinja"""
    result_name = 'results/:name.txt'.replace(":name", project_name)
    f = codecs.open(result_name, encoding='utf-8', mode='w')
    j2 = Environment(loader = FileSystemLoader(config['dir']), trim_blocks = config['trim_blocks'])
    for i in project_data:
        if(i["tipo"] != "General"):
            f.write(j2.get_template(config['jinja']['detallado']).render(
                proyecto = project_name,
                hallazgo = "*" + i["hallazgo"].encode('ascii', 'ignore') + "*", 
                vulnerabilidad = i["vulnerabilidad"].encode('ascii', 'ignore'),
                amenaza = i["amenaza"].encode('ascii', 'ignore'),
                riesgo = i["riesgo"].encode('ascii', 'ignore'),
                donde = i["donde"].encode('ascii', 'ignore'),
                solucion = i["solucion_efecto"].encode('ascii', 'ignore') + "\n"
                )
            )
        else:
            f.write(j2.get_template(config['jinja']['general']).render(
                proyecto = project_name,
                hallazgo = "*" + i["hallazgo"].encode('ascii', 'ignore') + "*", 
                vulnerabilidad = i["vulnerabilidad"].encode('ascii', 'ignore'),
                amenaza = i["amenaza"].encode('ascii', 'ignore'),
                riesgo = i["riesgo"].encode('ascii', 'ignore'),
                donde = i["donde"].encode('ascii', 'ignore'),
                solucion = i["solucion_efecto"].encode('ascii', 'ignore') + "\n"))
    return result_name

def write_to_cell(current_sheet, row, col, value, formula=False):
    """ Asigna un valor a una celda discriminando si es de tipo formula """
    temp_cell = current_sheet.cell(row=row, column=col)
    if not formula:
        temp_cell.set_explicit_value(value=value)
    else:
        temp_cell.set_explicit_value(value=value, data_type=temp_cell.TYPE_FORMULA)

def write_cell(current_sheet, row, col, value, formula=False):
    """ Asigna un valor a una celda discriminando si es de tipo formula """
    temp_cell = current_sheet.cell(row=row, column=col)
    temp_cell.value = value
    

def generate_doc_xls(project, data):
    """ Genera la documentacion xlsx de un proyecto """
    proy_type = "general"
    if "detallado" == data[0]["tipo"].lower():
        proy_type = "detallado"
    workbook_name = config['dir'] + config['xls'][proy_type]
    save_report_name = config['dir']+"/app/autodoc/results/" + project + ".xlsx"
    tech_report = load_workbook(filename=workbook_name)
    finding_sheet = tech_report["Hallazgos"]
    qc_sheet = tech_report["MatrizQC"]
    row = config["xls"]["first_row"]
    qc_row = config["xls"]["first_row"]
    data = util.ord_asc_by_criticidad(data)
    for i in data:
        # Columna Titulo (2)
        write_cell(finding_sheet, row, 2, i["hallazgo"])
        # Columna Vulnerabilidad (3)
        write_cell(finding_sheet, row, 3, i["vulnerabilidad"])
        # Columna Vulnerabilidad (4)
        write_cell(finding_sheet, row, 4, i["donde"])
        # Columna Requisitos (5)
        write_cell(finding_sheet, row, 5, i["requisitos"])
        # Columna Metricas (7)
        # Fix Autodoc formulario Bajo a Baja 
        vector_acceso = util.extract_metric(i["vector_acceso"])
        complejidad_acceso = util.extract_metric(i["complejidad_acceso"])
        # Fix Autodoc formulario Bajo a Baja 
        if complejidad_acceso == "Bajo":
            complejidad_acceso = "Baja"
        elif complejidad_acceso == "Alto":
            complejidad_acceso = "Alta"
        elif complejidad_acceso == "Medio":
            complejidad_acceso = "Media"
        autenticacion = util.extract_metric(i["autenticacion"])
        impacto_confidencialidad = util.extract_metric(i["impacto_confidencialidad"])
        impacto_integridad = util.extract_metric(i["impacto_integridad"])
        impacto_disponibilidad = util.extract_metric(i["impacto_disponibilidad"])
        explotabilidad = util.extract_metric(i["explotabilidad"])
        nivel_resolucion = util.extract_metric(i["nivel_resolucion"])
        nivel_confianza = util.extract_metric(i["nivel_confianza"])

        """
            TODO: 
                - Crear formulario para cambiar nombre de proyecto en bancolombia pedido
                - Extraer campos para documentar la QC
                - Verificar como insertar imagenes con ascii doctor
                - Verificar como descargar archivos de formstack
                - Diseñar formato para evidencias en formstack
                - Refactoring y linting
        """
        write_cell(finding_sheet, row, 7, vector_acceso)
        write_cell(finding_sheet, row+1, 7, complejidad_acceso)
        write_cell(finding_sheet, row+2, 7, autenticacion)
        write_cell(finding_sheet, row+3, 7, impacto_confidencialidad)
        write_cell(finding_sheet, row+4, 7, impacto_integridad)
        write_cell(finding_sheet, row+5, 7, impacto_disponibilidad)
        write_cell(finding_sheet, row+6, 7, explotabilidad)
        write_cell(finding_sheet, row+7, 7, nivel_resolucion)
        write_cell(finding_sheet, row+8, 7, nivel_confianza)
        # Columna Criticidad (8)
        write_cell(finding_sheet, row, 8, i["criticidad"])
        # Columna Cardinalidad (9)
        write_cell(finding_sheet, row, 9, int(i["cardinalidad"]))
        # Columna Evidencia (10)
        write_cell(finding_sheet, row, 10, "Evidencias/" + i["hallazgo"])
        # Columna Solucion efecto (11)
        write_cell(finding_sheet, row, 11, i["solucion_efecto"])
        # Columna Instructivo (12)
        write_cell(finding_sheet, row, 12, "Soluciones/"+ i["hallazgo"])
        # Columna Ids Requisitos (13)
        reqs = util.extract_reqs(i["requisitos"])
        write_cell(finding_sheet, row, 13 , ".*(" + reqs + ")")
        # Datos si es detallado
        if i["tipo"] == "Detallado":
            # Columna Amenaza (14)
            write_cell(finding_sheet, row, 14, i["amenaza"])
            # Columna Riesgo (15)
            write_cell(finding_sheet, row, 15, i["riesgo"])
            # Escribir en QC todos los campos


            # Columna Tipo de prueba (5)
            write_cell(qc_sheet, qc_row, 5, i["tipo_prueba"])
            # Columna Componente aplicativo (6)
            # Columna ID Requisitos (8)
            write_cell(qc_sheet, qc_row, 8, reqs) 
            # Columna Fabrica de testing (10)
            write_cell(qc_sheet, qc_row, 10, "Fluid")
            # Columna Detectador por (11)
            write_cell(qc_sheet, qc_row, 11, i["analista"])
            # Columna Detectado ciclo (12)
            # Columna Escenario (13)
            # FIX por modificacion de formstack
            try:
                escenario = i["escenario"]
            except KeyError: 
                escenario = ""
            write_cell(qc_sheet, qc_row, 13, escenario)
            # Columna Hallazgo (14)
            write_cell(qc_sheet, qc_row, 14, i["hallazgo"])
            # Columna Ambito (17)
            write_cell(qc_sheet, qc_row, 17, i["ambito"]) 
            # Columna Categoria Bancolombia (18) 
            # FIX por modificacion de formstack
            try:
                categoria = i["categoria"]
            except KeyError: 
                categoria = ""
            write_cell(qc_sheet, qc_row, 18, categoria)
            # Columna Amenaza (19)
            write_cell(qc_sheet, qc_row, 19, i["amenaza"])
            # Columna Tipo Reporte (20)
            write_cell(qc_sheet, qc_row, 20, "Hallazgo")
            # Columna Naturaleza (21)
            write_cell(qc_sheet, qc_row, 21, "Seguridad")
            # Columna Probabilidad (26)
            write_cell(qc_sheet, qc_row, 26, i["probabilidad"])
            # Columna Severidad (27)
            write_cell(qc_sheet, qc_row, 27, i["severidad"])
            qc_row = qc_row + 1
        row = row + 10
    tech_report.save(filename=save_report_name)