# -*- coding: utf-8 -*-
# Disabling this rule is necessary for include returns inside if-else structure
# pylint: disable-msg=no-else-return
# pylint: disable=too-many-lines
"""Views and services for FluidIntegrates."""

import os
import sys
import time
from datetime import datetime, timedelta

import boto3
import rollbar
import yaml
from django.conf import settings
from django.core.cache.backends.base import DEFAULT_TIMEOUT
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.decorators.cache import never_cache, cache_control
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from jose import jwt
from magic import Magic
from openpyxl import load_workbook, Workbook

from backend import util
from backend.domain import (
    finding as finding_domain, project as project_domain, user as user_domain)
from backend.domain.vulnerability import (
    group_specific, get_open_vuln_by_type, get_vulnerabilities_by_type
)
from backend.decorators import authenticate, authorize, cache_content
from backend.dal import (
    integrates_dal, finding as finding_dal, user as user_dal,
    project as project_dal
)
from backend.services import (
    has_access_to_project, has_access_to_finding, has_access_to_event
)
from backend.utils import reports

from __init__ import (
    FI_AWS_S3_ACCESS_KEY, FI_AWS_S3_SECRET_KEY, FI_AWS_S3_BUCKET
)

from app.documentator.pdf import CreatorPDF
from app.documentator.secure_pdf import SecurePDF
from app.documentator.all_vulns import generate_all_vulns_xlsx
from app.techdoc.it_report import ITReport

CACHE_TTL = getattr(settings, 'CACHE_TTL', DEFAULT_TIMEOUT)

CLIENT_S3 = boto3.client('s3',
                         aws_access_key_id=FI_AWS_S3_ACCESS_KEY,
                         aws_secret_access_key=FI_AWS_S3_SECRET_KEY)

BUCKET_S3 = FI_AWS_S3_BUCKET
BASE_URL = "https://fluidattacks.com/integrates"


@never_cache
def index(request):
    "Login view for unauthenticated users"
    parameters = {'debug': settings.DEBUG}
    return render(request, "index.html", parameters)


def error500(request):
    "Internal server error view"
    parameters = {}
    return render(request, "HTTP500.html", parameters)


def error401(request, _):
    "Unauthorized error view"
    parameters = {}
    return render(request, "HTTP401.html", parameters)


@csrf_exempt
@cache_control(private=True, max_age=3600)
@authenticate
def app(request):
    """ App view for authenticated users """
    try:
        parameters = {
            'debug': settings.DEBUG,
            'username': request.session['username']
        }
        response = render(request, 'app.html', parameters)
        token = jwt.encode(
            {
                'user_email': request.session['username'],
                'user_role': request.session['role'],
                'company': request.session['company'],
                'first_name': request.session['first_name'],
                'last_name': request.session['last_name'],
                'exp': datetime.utcnow() +
                timedelta(seconds=settings.SESSION_COOKIE_AGE)
            },
            algorithm='HS512',
            key=settings.JWT_SECRET,
        )
        response.set_cookie(
            key=settings.JWT_COOKIE_NAME,
            value=token,
            secure=True,
            httponly=True,
            max_age=settings.SESSION_COOKIE_AGE
        )
    except KeyError:
        rollbar.report_exc_info(sys.exc_info(), request)
        return redirect('/integrates/error500')
    return response


@csrf_exempt
@authenticate
def logout(request):
    "Close a user's active session"

    HttpResponse("<script>Intercom('shutdown');</script>")
    try:
        request.session.flush()
    except KeyError:
        rollbar.report_exc_info(sys.exc_info(), request)

    response = redirect("/integrates/index")
    response.delete_cookie(settings.JWT_COOKIE_NAME)
    return response


@cache_content
@never_cache
@csrf_exempt
@authorize(['analyst', 'customer', 'admin'])
def project_to_xls(request, lang, project):
    "Create the technical report"
    username = request.session['username'].split("@")[0]
    if project.strip() == "":
        rollbar.report_message(
            'Error: Empty fields in project', 'error', request)
        return util.response([], 'Empty fields', True)
    if not has_access_to_project(request.session['username'],
                                 project, request.session['role']):
        util.cloudwatch_log(
            request,
            'Security: Attempted to export project xls without permission')
        return util.response([], 'Access denied', True)
    if lang not in ["es", "en"]:
        rollbar.report_message('Error: Unsupported language', 'error', request)
        return util.response([], 'Unsupported language', True)
    findings = finding_domain.get_findings(
        project_domain.list_findings(project.lower()))
    if findings:
        findings = [cast_new_vulnerabilities(
            get_open_vuln_by_type(finding['findingId'], request), finding)
            for finding in findings]
    else:
        rollbar.report_message(
            'Project {} does not have findings in dynamo'.format(project),
            'warning',
            request)
        return util.response([], 'Empty fields', True)
    data = util.ord_asc_by_criticidad(findings)
    it_report = ITReport(project, data, username)
    filepath = it_report.result_filename
    reports.set_xlsx_password(filepath, time.strftime('%d%m%Y') + username)

    with open(filepath, 'rb') as document:
        response = HttpResponse(document.read())
        response['Content-Type'] = ('application/vnd.openxmlformats'
                                    '-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'inline;filename={}.xlsx'.format(
            project)
    return response


def validation_project_to_pdf(request, lang, doctype):
    if lang not in ["es", "en"]:
        rollbar.report_message('Error: Unsupported language', 'error', request)
        return util.response([], 'Unsupported language', True)
    if doctype not in ["tech", "executive"]:
        rollbar.report_message('Error: Unsupported doctype', 'error', request)
        return util.response([], 'Unsupported doctype', True)
    return None


@cache_content
@never_cache
@csrf_exempt
@authorize(['analyst', 'customer', 'admin'])
def project_to_pdf(request, lang, project, doctype):
    "Export a project to a PDF"
    assert project.strip()
    if not has_access_to_project(request.session['username'],
                                 project, request.session['role']):
        util.cloudwatch_log(request, 'Security: Attempted to export project'
                                     ' pdf without permission')
        return util.response([], 'Access denied', True)
    else:
        user = request.session['username'].split('@')[0]
        validator = validation_project_to_pdf(request, lang, doctype)
        if validator is not None:
            return validator
        findings = finding_domain.get_findings(
            project_domain.list_findings(project.lower()))
        findings = [cast_new_vulnerabilities(
            get_open_vuln_by_type(finding['findingId'], request), finding)
            for finding in findings]
        description = project_domain.get_description(project.lower())

        pdf_maker = CreatorPDF(lang, doctype)
        secure_pdf = SecurePDF()
        findings_ord = util.ord_asc_by_criticidad(findings)
        findings = pdf_evidences(findings_ord)
        report_filename = ''
        if doctype == 'tech':
            pdf_maker.tech(findings, project, description)
            report_filename = secure_pdf.create_full(user,
                                                     pdf_maker.out_name,
                                                     project)
        else:
            return HttpResponse(
                'Disabled report generation', content_type='text/html')
        if not os.path.isfile(report_filename):
            rollbar.report_message(
                'Couldn\'t generate pdf report', 'error', request)
            return HttpResponse(
                'Couldn\'t generate pdf report', content_type='text/html')
        with open(report_filename, 'rb') as document:
            response = HttpResponse(document.read(),
                                    content_type='application/pdf')
            response['Content-Disposition'] = \
                'inline;filename={}_IT.pdf'.format(project)
        return response


def pdf_evidences(findings):
    for finding in findings:
        folder_name = finding['projectName'] + '/' + finding['findingId']
        evidence = finding['evidence']
        evidence_set = [{
            'id': '{}/{}'.format(folder_name, evidence[ev_item]['url']),
            'explanation': evidence[ev_item]['description'].capitalize()
        } for ev_item in evidence if evidence[ev_item]['url'].endswith('.png')]

        if evidence_set:
            finding['evidence_set'] = evidence_set
            for evidence in evidence_set:
                CLIENT_S3.download_file(
                    BUCKET_S3,
                    evidence['id'],
                    '/usr/src/app/app/documentator/images/' +
                    evidence['id'].split('/')[2])
                evidence['name'] = 'image::../images/' + \
                    evidence['id'].split('/')[2] + '[align="center"]'

    return findings


def cast_new_vulnerabilities(finding_new, finding):
    """Cast values for new format."""
    if finding_new.get('openVulnerabilities') >= 0:
        finding['openVulnerabilities'] = \
            str(finding_new.get('openVulnerabilities'))
    else:
        # This finding does not have open vulnerabilities
        pass
    where = '-'
    if finding_new.get('portsVulns'):
        finding['portsVulns'] = \
            group_specific(finding_new.get('portsVulns'), 'ports')
        where = format_where(where, finding['portsVulns'])
    else:
        # This finding does not have ports vulnerabilities
        pass
    if finding_new.get('linesVulns'):
        finding['linesVulns'] = \
            group_specific(finding_new.get('linesVulns'), 'lines')
        where = format_where(where, finding['linesVulns'])
    else:
        # This finding does not have lines vulnerabilities
        pass
    if finding_new.get('inputsVulns'):
        finding['inputsVulns'] = \
            group_specific(finding_new.get('inputsVulns'), 'inputs')
        where = format_where(where, finding['inputsVulns'])
    else:
        # This finding does not have inputs vulnerabilities
        pass
    finding['where'] = where
    return finding


def format_where(where, vulnerabilities):
    """Formate where field with new vulnerabilities."""
    for vuln in vulnerabilities:
        where = '{where!s}{vuln_where!s} ({vuln_specific!s})\n'\
                .format(where=where,
                        vuln_where=vuln.get('where'),
                        vuln_specific=vuln.get('specific'))
    return where


def format_release_date(finding):
    primary_keys = ["finding_id", finding["id"]]
    table_name = "FI_findings"
    finding_dynamo = integrates_dal.get_data_dynamo(
        table_name, primary_keys[0], primary_keys[1])
    if finding_dynamo:
        if finding_dynamo[0].get("releaseDate"):
            finding["releaseDate"] = finding_dynamo[0].get("releaseDate")
        if finding_dynamo[0].get("lastVulnerability"):
            finding["lastVulnerability"] = \
                finding_dynamo[0].get("lastVulnerability")
    if finding.get("releaseDate"):
        final_date = util.calculate_datediff_since(finding["releaseDate"])
        finding['edad'] = final_date.days
        final_vuln_date = util.calculate_datediff_since(finding["lastVulnerability"])
        finding['lastVulnerability'] = final_vuln_date.days
    else:
        finding['lastVulnerability'] = '-'
    return finding


@cache_content
@never_cache
@csrf_exempt
@authorize(['analyst', 'customer', 'admin'])
def get_evidence(request, project, evidence_type, findingid, fileid):
    username = request.session['username']
    role = request.session['role']
    if (evidence_type in ['drafts', 'findings']
        and has_access_to_finding(username, findingid, role)) \
            or (evidence_type == 'events'
                and has_access_to_event(username, findingid, role)):
        if fileid is None:
            rollbar.report_message('Error: Missing evidence image ID',
                                   'error', request)
            return HttpResponse("Error - Unsent image ID",
                                content_type="text/html")
        key_list = key_existing_list(f'{project.lower()}/{findingid}/{fileid}')
        if key_list:
            for k in key_list:
                start = k.find(findingid) + len(findingid)
                localfile = "/tmp" + k[start:]
                ext = {'.png': '.tmp', '.gif': '.tmp'}
                localtmp = util.replace_all(localfile, ext)
                CLIENT_S3.download_file(BUCKET_S3, k, localtmp)
                return retrieve_image(request, localtmp)
        else:
            return util.response([], 'Access denied or evidence not found', True)
    else:
        util.cloudwatch_log(
            request,
            'Security: Attempted to retrieve evidence without permission')
        return util.response([], 'Access denied or evidence not found', True)


def retrieve_image(request, img_file):
    if util.assert_file_mime(img_file, ["image/png", "image/jpeg",
                                        "image/gif"]):
        with open(img_file, "rb") as file_obj:
            mime = Magic(mime=True)
            mime_type = mime.from_file(img_file)
            return HttpResponse(file_obj.read(), content_type=mime_type)
    else:
        rollbar.report_message('Error: Invalid evidence image format',
                               'error', request)
        return HttpResponse("Error: Invalid evidence image format",
                            content_type="text/html")


def key_existing_list(key):
    """return the key's list if it exist, else list empty"""
    return util.list_s3_objects(CLIENT_S3, BUCKET_S3, key)


def delete_project(project):
    """Delete project information."""
    project = project.lower()
    are_users_removed = remove_all_users_access(project)
    are_findings_masked = [
        finding_domain.mask_finding(finding_id)
        for finding_id in project_domain.list_findings(project)]
    update_project_state_db = project_domain.update(project, {'project_status': 'FINISHED'})
    is_project_deleted = all([
        are_findings_masked, are_users_removed, update_project_state_db])
    util.invalidate_cache(project)

    return is_project_deleted


def remove_all_users_access(project):
    """Remove user access to project."""
    user_active = project_domain.get_users(project)
    user_suspended = project_domain.get_users(project, active=False)
    all_users = user_active + user_suspended
    are_users_removed = True
    for user in all_users:
        is_user_removed = remove_user_access(project, user)
        if is_user_removed:
            are_users_removed = True
        else:
            are_users_removed = False
            break
    return are_users_removed


def remove_user_access(project, user_email):
    """Remove user access to project."""
    integrates_dal.remove_role_to_project_dynamo(
        project, user_email, 'customeradmin')
    return project_dal.remove_access(user_email, project)


@cache_content
@never_cache
@csrf_exempt
@require_http_methods(["GET"])
@authorize(['analyst', 'admin'])
def download_vulnerabilities(request, findingid):
    """Download a file with all the vulnerabilities."""
    if not has_access_to_finding(request.session['username'], findingid,
                                 request.session['role']):
        util.cloudwatch_log(request,
                            'Security: \
Attempted to retrieve vulnerabilities without permission')
        return util.response([], 'Access denied', True)
    else:
        finding = get_vulnerabilities_by_type(findingid)
        data_yml = {}
        vuln_types = {'ports': dict, 'lines': dict, 'inputs': dict}
        if finding:
            for vuln_key, cast_fuction in list(vuln_types.items()):
                if finding.get(vuln_key):
                    data_yml[vuln_key] = list(map(cast_fuction, list(finding.get(vuln_key))))
                else:
                    # This finding does not have this type of vulnerabilities
                    pass
        else:
            # This finding does not have new vulnerabilities
            pass
        project = finding_domain.get_finding(findingid)['projectName']
        file_name = '/tmp/{project}-{finding_id}.yaml'.format(
            finding_id=findingid, project=project)
        stream = open(file_name, 'w')
        yaml.safe_dump(data_yml, stream, default_flow_style=False)
        try:
            with open(file_name, 'rb') as file_obj:
                response = HttpResponse(file_obj.read(), content_type='text/x-yaml')
                response['Content-Disposition'] = \
                    'attachment; filename="{project}-{finding_id}.yaml"'.format(
                        finding_id=findingid, project=project)
                return response
        except IOError:
            rollbar.report_message('Error: Invalid vulnerabilities file format', 'error', request)
            return util.response([], 'Invalid vulnerabilities file format', True)


@never_cache
@require_http_methods(["GET"])
# pylint: disable=too-many-locals
def generate_complete_report(request):
    user_data = util.get_jwt_content(request)
    projects = user_domain.get_projects(user_data['user_email'])
    book = load_workbook('/usr/src/app/app/techdoc/templates/COMPLETE.xlsx')
    sheet = book.active

    project_col = 1
    finding_col = 2
    vuln_where_col = 3
    vuln_specific_col = 4
    treatment_col = 5
    treatment_mgr_col = 6
    row_offset = 2

    row_index = row_offset
    for project in projects:
        findings = integrates_dal.get_findings_released_dynamo(
            project, 'finding_id, finding, treatment')
        for finding in findings:
            vulns = finding_dal.get_vulnerabilities(finding['finding_id'])
            for vuln in vulns:
                sheet.cell(row_index, vuln_where_col, vuln['where'])
                sheet.cell(row_index, vuln_specific_col, vuln['specific'])

                sheet.cell(row_index, project_col, project.upper())
                sheet.cell(row_index, finding_col, '{name!s} (#{id!s})'.format(
                           name=finding['finding'].encode('utf-8'),
                           id=finding['finding_id']))
                sheet.cell(row_index, treatment_col, finding['treatment'])
                sheet.cell(row_index, treatment_mgr_col,
                           vuln.get('treatment_manager', 'Unassigned'))

                row_index += 1

    username = user_data['user_email'].split('@')[0].encode('utf8', 'ignore')
    filename = 'complete_report.xlsx'
    filepath = '/tmp/{username}-{filename}'.format(filename=filename,
                                                   username=username)
    book.save(filepath)

    with open(filepath, 'rb') as document:
        response = HttpResponse(document.read())
        response['Content-Type'] = 'application/vnd.openxmlformats\
                        -officedocument.spreadsheetml.sheet'
        response['Content-Disposition'] = 'inline;filename={filename}'.format(
            filename=filename)
    return response


@cache_content
@never_cache
@authorize(['admin'])
def export_all_vulnerabilities(request):
    user_data = util.get_jwt_content(request)
    filepath = generate_all_vulns_xlsx(user_data['user_email'])
    filename = os.path.basename(filepath)
    with open(filepath, 'rb') as document:
        response = HttpResponse(document.read())
        response['Content-Type'] = 'application/vnd.openxmlformats\
                        -officedocument.spreadsheetml.sheet'
        response['Content-Disposition'] = 'inline;filename={filename}'.format(
            filename=filename)
    return response


@cache_content
@never_cache
@authorize(['admin'])
def export_users(request):
    user_data = util.get_jwt_content(request)
    book = Workbook()
    sheet = book.active
    sheet.append(['full_name', 'user_email'])
    row_index = 2

    unique_users = []
    for user in user_dal.get_platform_users():
        user_email = user['user_email'].lower()
        if user_email not in unique_users:
            unique_users.append(user_email)

            name_attrs = user_domain.get_attributes(
                user_email, ['first_name', 'last_name'])
            full_name = ' '.join(list(name_attrs.values()))

            sheet.cell(row_index, 1, full_name)
            sheet.cell(row_index, 2, user_email)
            row_index += 1

    username = user_data['user_email'].split('@')[0].encode('utf8', 'ignore')
    filepath = f'/tmp/{username}-users.xlsx'
    filename = os.path.basename(filepath)
    book.save(filepath)

    with open(filepath, 'rb') as document:
        response = HttpResponse(document.read())
        response['Content-Type'] = 'application/vnd.openxmlformats\
                        -officedocument.spreadsheetml.sheet'
        response['Content-Disposition'] = f'inline;filename={filename}'
    return response
