# -*- coding: utf-8 -*-
""" Export all vulnerabilities """


import hashlib
from openpyxl import Workbook

from backend.dal import finding as finding_dal, project as project_dal

from __init__ import FI_TEST_PROJECTS

TEST_PROJECTS = FI_TEST_PROJECTS.split(',')

COLUMNS_FINS = [
    'project_name',
    'finding_id',
    'finding',
    'finding_type',
    'attack_vector',
    'attack_complexity',
    'user_interaction',
    'severity_scope',
    'confidentiality_impact',
    'integrity_impact',
    'availability_impact',
    'exploitability',
    'remediation_level',
    'report_confidence',
    'cvss_basescore',
    'cvss_temporal',
    'actor',
    'cwe',
    'scenario'
]

COLUMNS_VULNS = [
    'vuln_type',
    'report_date',
    'analyst',
    'treatment',
    'specific',
    'closing_date'
]

MASKED_COLUMNS = [
    'finding_id',
    'project_name',
]


def _hash_cell(cell):
    return hashlib.sha256(cell.encode()).hexdigest()[-5:]


def _mask_finding(finding):
    for masked_cell in MASKED_COLUMNS:
        finding[masked_cell] = _hash_cell(finding[masked_cell])
    return finding


def _get_reporter_analyst(opening_state, vuln, finding):
    analyst = opening_state.get('analyst')
    if not analyst:
        analyst = vuln.get('analyst')
        if not analyst:
            finding.get('analyst')
    if analyst:
        analyst = _hash_cell(analyst)
    else:
        analyst = ''
    return analyst


def _format_specific(vuln):
    specific = ''
    if vuln.get('specific') != 'Masked':
        if vuln.get('vuln_type') == 'lines':
            file_ext = vuln.get('where').split('/')[-1].split('.')[-1]
            if (file_ext.isalnum()
                    and not file_ext.isdigit()
                    and file_ext != 'Masked'):
                specific = file_ext.lower()
            else:
                specific = ''
        elif vuln.get('vuln_type') == 'ports':
            specific = vuln.get('specific')
        else:
            specific = ''
    else:
        specific = ''

    return specific


def _format_vuln(vuln, finding):
    if not vuln.get('treatment'):
        if finding.get('TREATMENT'):
            vuln['treatment'] = finding.get('TREATMENT')
        else:
            vuln['treatment'] = 'NEW'

    historic_state = vuln.get('historic_state')
    last_state = historic_state[-1]
    opening_state = historic_state[0]

    if last_state.get('state') == 'closed':
        vuln['treatment'] = 'CLOSED'
        vuln['closing_date'] = last_state.get('date')

    vuln['specific'] = _format_specific(vuln)
    vuln['report_date'] = opening_state.get('date')
    vuln['analyst'] = _get_reporter_analyst(opening_state, vuln, finding)
    return vuln


def fill_sheet(sheet, finding_row, vuln_row, row_index):
    for col, col_name in enumerate(COLUMNS_FINS, 1):
        sheet.cell(row_index, col, finding_row.get(col_name, ''))
    for col, col_name in enumerate(COLUMNS_VULNS, len(COLUMNS_FINS) + 1):
        sheet.cell(row_index, col, vuln_row.get(col_name, ''))


def generate_all_vulns_xlsx(user_email):
    projects = project_dal.get_all()
    book = Workbook()
    sheet = book.active
    sheet.append(COLUMNS_FINS + COLUMNS_VULNS)
    row_index = 2
    for project in projects:
        if project not in TEST_PROJECTS:
            findings = project_dal.get_released_findings(project)
        else:
            findings = []
        for finding in findings:
            vulns = finding_dal.get_vulnerabilities(finding['finding_id'])
            finding_row = _mask_finding(finding)
            for vuln in vulns:
                vuln_row = _format_vuln(vuln, finding_row)
                fill_sheet(sheet, finding_row, vuln_row, row_index)
                row_index += 1

    username = user_email.split('@')[0].encode('utf8', 'ignore')
    filepath = '/tmp/{username}-all_vulns.xlsx'.format(username=username)
    book.save(filepath)
    return filepath
